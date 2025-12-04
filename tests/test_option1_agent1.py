"""
Test Option 1: Agent 1 (Email Extractor) standalone
"""
from src.ui.agent_runner import _prompt_agent1_parameters, _execute_agent1

print("="*70)
print("Testing Option 1: Agent 1 (Email Extractor) - Standalone")
print("="*70)

# Use search term that previously worked (8 emails found)
email_subject = "AI Development Expert course - Homework - L"
sender_email = None
max_emails = 20

print(f"\nEmail Subject: {email_subject}")
print(f"Sender Email: {sender_email}")
print(f"Max Emails: {max_emails}\n")

# Build query using current logic (lines 168-189)
words = email_subject.split()
if len(words) <= 3:
    query = f'subject:{email_subject}'
else:
    key_words = ' '.join(words[:3])
    query = f'subject:{key_words}'

print(f"Expected Gmail Query: {query}")
print(f"Expected Post-Filter: '{email_subject}' as substring (case-insensitive)")
print("\n" + "="*70)

params = {
    'email_subject': email_subject,
    'sender_email': sender_email,
    'max_emails': max_emails,
    'query': query
}

print("Executing Agent 1...")
success = _execute_agent1(params)

if success:
    import openpyxl
    from src.utils.paths import get_excel_file

    excel_path = get_excel_file('Excel1.xlsx')
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        row_count = ws.max_row - 1
        print(f"\n{'='*70}")
        print(f"RESULT: {row_count} emails found")
        print(f"{'='*70}\n")

        if row_count > 0:
            print("Email subjects found:")
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                email_id, received_time, subject, sender, github_url, thread_id, status = row
                print(f"  {i}. {subject}")

            if row_count == 8:
                print(f"\nSUCCESS: Found expected 8 emails")
            else:
                print(f"\nWARNING: Expected 8 emails, got {row_count}")
        else:
            print("FAILURE: No emails found")
    else:
        print("FAILURE: Excel1.xlsx not created")
else:
    print("FAILURE: Agent 1 execution failed")
