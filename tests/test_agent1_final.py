"""
Final test with new query building logic
"""

email_subject = 'AI Development Expert course - Homework - L'

# Build query (new logic from agent_runner.py)
words = email_subject.split()
if len(words) > 3:
    key_words = ' '.join(words[:3])
    query = f'subject:{key_words}'
else:
    query = f'subject:{email_subject}'

print(f"User Search: {email_subject}")
print(f"Gmail Query: {query}")
print(f"Number of words: {len(words)}")
print()

# Test with Agent 1
params = {
    'email_subject': email_subject,
    'sender_email': None,
    'max_emails': 20,
    'query': query
}

print("Running Agent 1...")
print("="*70)

from src.ui.agent_runner import _execute_agent1

success = _execute_agent1(params)

if success:
    import openpyxl
    from src.utils.paths import get_excel_file

    excel_path = get_excel_file('Excel1.xlsx')
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        row_count = ws.max_row - 1
        print(f"\n[+] RESULT: {row_count} emails found")

        if row_count > 0:
            print("\nEmail subjects:")
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                email_id, received_time, subject, sender, github_url, thread_id, status = row
                print(f"{i}. {subject}")
