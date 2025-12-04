"""
Quick test to verify Agent 2 can read the new Excel1.xlsx format
"""
import openpyxl
from src.utils.paths import get_excel_file

# Test reading Excel1.xlsx with new format (7 columns)
excel1_path = get_excel_file('Excel1.xlsx')

if not excel1_path.exists():
    print("[!] Excel1.xlsx not found. Run Agent 1 first.")
else:
    wb = openpyxl.load_workbook(excel1_path)
    ws = wb.active

    print("Excel1.xlsx Column Headers:")
    headers = [cell.value for cell in ws[1]]
    print(f"  {headers}")
    print(f"  Total columns: {len(headers)}")
    print()

    print("Testing row unpacking (new format - 7 columns):")
    try:
        ready_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # New format: 7 columns (removed hashed_email)
            email_id, received_time, subject, sender, github_url, thread_id, status = row
            if status == "Ready":
                ready_rows.append({
                    'email_id': email_id,
                    'github_url': github_url
                })

        print(f"[+] Successfully unpacked {len(ready_rows)} ready rows")
        print(f"[+] Format is correct!")

        # Show first 3
        for i, row_data in enumerate(ready_rows[:3], 1):
            print(f"  {i}. {row_data['github_url']}")

    except ValueError as e:
        print(f"[!] Error: {e}")
        print("[!] Format mismatch - check column count")
