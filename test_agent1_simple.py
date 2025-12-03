"""
Simple test to check Gmail query building
"""

email_subject = 'AI Development Expert course - Homework - L'

# Build query (same logic as agent_runner.py lines 170-177)
query_parts = []
if email_subject:
    # Use quotes if search term contains spaces
    if ' ' in email_subject:
        query_parts.append(f'subject:"{email_subject}"')
    else:
        query_parts.append(f'subject:{email_subject}')

query = ' '.join(query_parts) if query_parts else ''

print(f"Email Subject: {email_subject}")
print(f"Gmail Query: {query}")
print()

# Now test with Agent 1
print("Testing with Agent 1...")
print("="*70)

params = {
    'email_subject': email_subject,
    'sender_email': None,
    'max_emails': 20,
    'query': query
}

from src.ui.agent_runner import _execute_agent1

success = _execute_agent1(params)

if success:
    import openpyxl
    from src.utils.paths import get_excel_file

    excel_path = get_excel_file('Excel1.xlsx')
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        row_count = ws.max_row - 1  # Subtract header
        print(f"\n[+] Found {row_count} emails")

        if row_count > 0:
            print("\nEmail subjects:")
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                email_id, received_time, subject, sender, github_url, thread_id, status = row
                print(f"{i}. {subject}")
