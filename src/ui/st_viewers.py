"""
Streamlit Viewer Components - Data viewers and system status displays

Author: Hadar Wayn
Date: December 2025
"""

import streamlit as st
import pandas as pd
from pathlib import Path
from typing import Dict, Any
import openpyxl
from datetime import datetime


def render_excel_viewer(excel_path: Path, title: str):
    """Render an Excel file viewer with interactive table"""
    if not excel_path.exists():
        st.warning(f"📄 {title} not found. Run the corresponding agent first.")
        return
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
            else:
                data.append(row)
        if not data:
            st.info(f"📊 {title} is empty. No data to display.")
            return
        df = pd.DataFrame(data, columns=headers)
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Rows", len(df))
        with col2:
            ready_count = len(df[df['status'] == 'Ready']) if 'status' in df.columns else 0
            st.metric("Ready", ready_count)
        with col3:
            if 'grade' in df.columns:
                avg_grade = df['grade'].mean()
                st.metric("Avg Grade", f"{avg_grade:.1f}")
        st.dataframe(
            df,
            use_container_width=True,
            height=400
        )
        csv = df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download CSV",
            data=csv,
            file_name=f"{excel_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
    except Exception as e:
        st.error(f"Error reading {title}: {str(e)}")


def render_system_status() -> Dict[str, Any]:
    """Render system status widget"""
    from src.utils.paths import get_excel_file, get_temp_dir
    status_data = {}
    excel_files = {
        'Excel1.xlsx': get_excel_file('Excel1.xlsx'),
        'Excel2.xlsx': get_excel_file('Excel2.xlsx'),
        'Excel3.xlsx': get_excel_file('Excel3.xlsx')
    }
    status_data['excel_files'] = {}
    for name, path in excel_files.items():
        if path.exists():
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            status_data['excel_files'][name] = {
                'exists': True,
                'rows': ws.max_row - 1,
                'modified': datetime.fromtimestamp(path.stat().st_mtime)
            }
        else:
            status_data['excel_files'][name] = {'exists': False}
    temp_dir = get_temp_dir()
    if temp_dir.exists():
        repo_count = len([d for d in temp_dir.iterdir() if d.is_dir() and d.name != '.gitkeep'])
        status_data['repos'] = repo_count
    else:
        status_data['repos'] = 0
    return status_data


def render_status_cards(status_data: Dict[str, Any]):
    """Render status cards showing system state"""
    st.markdown("### 📊 System Status")
    cols = st.columns(4)
    excel_exists = [v['exists'] for v in status_data['excel_files'].values()]
    completed_agents = sum(excel_exists)
    with cols[0]:
        st.metric(
            "Agents Completed",
            f"{completed_agents}/3",
            delta=None
        )
    excel1_data = status_data['excel_files'].get('Excel1.xlsx', {})
    with cols[1]:
        if excel1_data.get('exists'):
            st.metric("Emails Extracted", excel1_data['rows'])
        else:
            st.metric("Emails Extracted", "—")
    excel2_data = status_data['excel_files'].get('Excel2.xlsx', {})
    with cols[2]:
        if excel2_data.get('exists'):
            st.metric("Repos Analyzed", excel2_data['rows'])
        else:
            st.metric("Repos Analyzed", "—")
    with cols[3]:
        st.metric("Cloned Repos", status_data['repos'])
