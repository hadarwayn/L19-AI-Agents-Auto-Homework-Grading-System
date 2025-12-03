"""
Streamlit UI Components for L19 Grading System
Reusable UI components for the modern web interface
"""

import streamlit as st
import pandas as pd
from pathlib import Path
from typing import Optional, Dict, Any
import openpyxl
from datetime import datetime


def render_header():
    """Render the application header with logo and title"""
    st.set_page_config(
        page_title="L19 AI Grading System",
        page_icon="🤖",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    st.markdown("""
        <style>
        .main-header {
            text-align: center;
            padding: 2rem 0;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 10px;
            margin-bottom: 2rem;
        }
        .agent-card {
            padding: 1.5rem;
            border-radius: 10px;
            border: 2px solid #e0e0e0;
            margin: 1rem 0;
            transition: all 0.3s;
        }
        .agent-card:hover {
            border-color: #667eea;
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.2);
        }
        .status-badge {
            padding: 0.5rem 1rem;
            border-radius: 20px;
            font-weight: bold;
            display: inline-block;
        }
        .status-ready {
            background-color: #4caf50;
            color: white;
        }
        .status-pending {
            background-color: #ff9800;
            color: white;
        }
        .status-error {
            background-color: #f44336;
            color: white;
        }
        </style>
    """, unsafe_allow_html=True)

    st.markdown("""
        <div class="main-header">
            <h1>🤖 L19 AI Agents Auto Homework Grading System</h1>
            <p style="font-size: 1.2rem; margin-top: 0.5rem;">
                Modern Web Interface • Automated Grading Pipeline • AI-Powered Feedback
            </p>
        </div>
    """, unsafe_allow_html=True)


def render_agent_card(
    agent_number: int,
    agent_name: str,
    description: str,
    icon: str,
    button_label: str,
    on_click: callable
) -> bool:
    """Render an agent card with clickable button"""
    with st.container():
        col1, col2 = st.columns([3, 1])

        with col1:
            st.markdown(f"### {icon} Agent {agent_number}: {agent_name}")
            st.markdown(f"<p style='color: #666;'>{description}</p>", unsafe_allow_html=True)

        with col2:
            return st.button(
                button_label,
                key=f"agent_{agent_number}_btn",
                use_container_width=True,
                type="primary"
            )


def render_agent1_form() -> Optional[Dict[str, Any]]:
    """Render Agent 1 parameter form with input fields"""
    st.markdown("### 📧 Email Search Parameters")

    with st.form("agent1_form"):
        col1, col2 = st.columns(2)

        with col1:
            email_subject = st.text_input(
                "Email Subject",
                placeholder="e.g., GradingL19 or Homework",
                help="Search term for email subject (case-insensitive substring match)"
            )

            sender_email = st.text_input(
                "Sender Email (Optional)",
                placeholder="student@example.com",
                help="Filter by sender email address (leave empty for all senders)"
            )

        with col2:
            max_emails = st.slider(
                "Maximum Emails",
                min_value=1,
                max_value=100,
                value=20,
                help="Maximum number of emails to process (1-100)"
            )

            st.markdown(f"""
            <div style='padding: 1rem; background-color: #f0f2f6; border-radius: 8px; margin-top: 1rem;'>
                <strong>📊 Search Preview:</strong><br>
                • Subject: <code>{email_subject if email_subject else '(all)'}</code><br>
                • Sender: <code>{sender_email if sender_email else '(all)'}</code><br>
                • Max: <code>{max_emails}</code> emails
            </div>
            """, unsafe_allow_html=True)

        submitted = st.form_submit_button("🔍 Search & Extract Emails", use_container_width=True)

        if submitted:
            # Build Gmail search query (same logic as agent_runner.py)
            query_parts = []

            if email_subject:
                # Gmail search WITHOUT quotes works best - it finds emails containing all words
                # Python post-filter ensures exact substring match (case-insensitive)
                # Use first 3-4 words for broad Gmail search, post-filter ensures accuracy
                words = email_subject.split()

                if len(words) <= 3:
                    # Short search (1-3 words): use all words
                    query_parts.append(f'subject:{email_subject}')
                else:
                    # Longer search (4+ words): use first 3 words for broad Gmail match
                    # Post-filter will ensure full phrase is in subject
                    key_words = ' '.join(words[:3])
                    query_parts.append(f'subject:{key_words}')

            if sender_email:
                query_parts.append(f'from:{sender_email}')

            # Combine query parts (search all emails if no filters)
            query = ' '.join(query_parts) if query_parts else ''

            return {
                'email_subject': email_subject if email_subject else None,
                'sender_email': sender_email if sender_email else None,
                'max_emails': max_emails,
                'query': query
            }

    return None


def render_progress_tracker(agent_name: str, status: str = "running"):
    """Render a progress tracker for agent execution"""
    if status == "running":
        st.info(f"⏳ {agent_name} is running...")
        progress_bar = st.progress(0)
        return progress_bar
    elif status == "success":
        st.success(f"✅ {agent_name} completed successfully!")
    elif status == "error":
        st.error(f"❌ {agent_name} failed!")


def render_excel_viewer(excel_path: Path, title: str):
    """Render an Excel file viewer with interactive table"""
    if not excel_path.exists():
        st.warning(f"📄 {title} not found. Run the corresponding agent first.")
        return

    try:
        # Read Excel file
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Convert to pandas DataFrame
        data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
            else:
                data.append(row)

        if not data:
            st.info(f"📊 {title} is empty. No data to display.")
            return

        df = pd.DataFrame(data, columns=headers)

        # Display metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Rows", len(df))
        with col2:
            ready_count = len(df[df['status'] == 'Ready']) if 'status' in df.columns else 0
            st.metric("Ready", ready_count)
        with col3:
            if 'grade' in df.columns:
                avg_grade = df['grade'].mean()
                st.metric("Avg Grade", f"{avg_grade:.1f}")

        # Display table with search and filter
        st.dataframe(
            df,
            use_container_width=True,
            height=400
        )

        # Download button
        csv = df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download CSV",
            data=csv,
            file_name=f"{excel_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )

    except Exception as e:
        st.error(f"Error reading {title}: {str(e)}")


def render_system_status() -> Dict[str, Any]:
    """Render system status widget"""
    from src.utils.paths import get_excel_file, get_temp_dir

    status_data = {}

    # Check Excel files
    excel_files = {
        'Excel1.xlsx': get_excel_file('Excel1.xlsx'),
        'Excel2.xlsx': get_excel_file('Excel2.xlsx'),
        'Excel3.xlsx': get_excel_file('Excel3.xlsx')
    }

    status_data['excel_files'] = {}
    for name, path in excel_files.items():
        if path.exists():
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            status_data['excel_files'][name] = {
                'exists': True,
                'rows': ws.max_row - 1,
                'modified': datetime.fromtimestamp(path.stat().st_mtime)
            }
        else:
            status_data['excel_files'][name] = {'exists': False}

    # Check temp repos
    temp_dir = get_temp_dir()
    if temp_dir.exists():
        repo_count = len([d for d in temp_dir.iterdir() if d.is_dir() and d.name != '.gitkeep'])
        status_data['repos'] = repo_count
    else:
        status_data['repos'] = 0

    return status_data


def render_status_cards(status_data: Dict[str, Any]):
    """Render status cards showing system state"""
    st.markdown("### 📊 System Status")

    cols = st.columns(4)

    # Excel files status
    excel_exists = [v['exists'] for v in status_data['excel_files'].values()]
    completed_agents = sum(excel_exists)

    with cols[0]:
        st.metric(
            "Agents Completed",
            f"{completed_agents}/3",
            delta=None
        )

    # Excel1 status
    excel1_data = status_data['excel_files'].get('Excel1.xlsx', {})
    with cols[1]:
        if excel1_data.get('exists'):
            st.metric("Emails Extracted", excel1_data['rows'])
        else:
            st.metric("Emails Extracted", "—")

    # Excel2 status
    excel2_data = status_data['excel_files'].get('Excel2.xlsx', {})
    with cols[2]:
        if excel2_data.get('exists'):
            st.metric("Repos Analyzed", excel2_data['rows'])
        else:
            st.metric("Repos Analyzed", "—")

    # Temp repos
    with cols[3]:
        st.metric("Cloned Repos", status_data['repos'])
