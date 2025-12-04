"""
Agent 4 Executor - Draft Email Creation

Creates Gmail draft emails with personalized feedback.

Author: Hadar Wayn
Date: December 2025
"""

import openpyxl
from pathlib import Path
from email.mime.text import MIMEText
import base64
from rich.console import Console

from .gmail_auth import get_gmail_service
from ...utils.paths import get_excel_file

console = Console()


def execute_agent4() -> bool:
    """Execute Agent 4 draft email creation logic"""
    try:
        excel3_path = get_excel_file('Excel3.xlsx')
        excel1_path = get_excel_file('Excel1.xlsx')
        if not excel3_path.exists():
            console.print("[red][!] Excel3.xlsx not found. Run Agent 3 first.[/red]\n")
            return False
        if not excel1_path.exists():
            console.print("[red][!] Excel1.xlsx not found. Run Agent 1 first.[/red]\n")
            return False
        console.print(f"[*] Reading Excel3: {excel3_path}")
        console.print(f"[*] Reading Excel1: {excel1_path}\n")

        feedback_data = _load_feedback_data(excel3_path)
        console.print(f"[+] Found {len(feedback_data)} submissions with feedback\n")
        if len(feedback_data) == 0:
            console.print("[!] No submissions with feedback to process.\n")
            return True
        email_metadata = _load_email_metadata(excel1_path)
        student_names = _load_student_mappings()
        service = get_gmail_service(console)
        drafts_created, drafts_failed = _create_gmail_drafts(
            service, feedback_data, email_metadata, student_names
        )
        _print_summary(drafts_created, drafts_failed)
        return True

    except Exception as e:
        console.print(f"\n[red][!] Agent 4 execution failed: {e}[/red]\n")
        import traceback
        console.print(f"[dim]{traceback.format_exc()}[/dim]")
        return False

def _load_feedback_data(excel3_path: Path) -> dict:
    """Load feedback data from Excel3"""
    wb = openpyxl.load_workbook(excel3_path)
    ws = wb.active
    feedback_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        email_id, grade, category, persona, response, attempts, status = row
        if status == "Ready" and response:
            feedback_data[email_id] = {'grade': grade, 'response': response}
    return feedback_data

def _load_email_metadata(excel1_path: Path) -> dict:
    """Load email metadata from Excel1"""
    wb = openpyxl.load_workbook(excel1_path)
    ws = wb.active
    email_metadata = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        email_id, received_time, subject, sender, github_url, thread_id, status = row
        email_metadata[email_id] = {
            'sender': sender,
            'subject': subject,
            'thread_id': thread_id,
            'github_url': github_url
        }
    return email_metadata

def _load_student_mappings() -> dict:
    """Load student name mappings (optional)"""
    student_names = {}
    mapping_path = Path('data/students_mapping.xlsx')
    if mapping_path.exists():
        try:
            wb = openpyxl.load_workbook(mapping_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                email, name = row
                if email and name:
                    student_names[email.lower().strip()] = name
            console.print(f"[+] Loaded {len(student_names)} student name mappings\n")
        except Exception as e:
            console.print(f"[!] Could not load student mappings: {e}\n")
    return student_names

def _create_gmail_drafts(service, feedback_data: dict, email_metadata: dict, student_names: dict) -> tuple:
    """Create Gmail drafts for all submissions"""
    console.print("[*] Creating Gmail drafts...\n")
    drafts_created = 0
    drafts_failed = 0

    for email_id, feedback in feedback_data.items():
        metadata = email_metadata.get(email_id)
        if not metadata:
            console.print(f"  [!] No metadata for {email_id[:8]}, skipping")
            drafts_failed += 1
            continue

        sender_email = metadata['sender']
        original_subject = metadata['subject']
        thread_id = metadata['thread_id']
        github_url = metadata['github_url']
        grade = feedback['grade']
        response_text = feedback['response']

        student_name = student_names.get(sender_email.lower(), sender_email.split('@')[0])
        email_body = f"Hi {student_name}!\n\n{response_text}\n\nYour repository reviewed: {github_url}\nYour grade: {grade}/100\n\nBest regards,\nInstructor\n"
        message = MIMEText(email_body)
        message['to'] = sender_email
        message['subject'] = f"Re: {original_subject}"
        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
        try:
            draft = {'message': {'raw': raw_message, 'threadId': thread_id}}
            service.users().drafts().create(userId='me', body=draft).execute()
            console.print(f"  [+] Draft created for {student_name} ({sender_email})")
            drafts_created += 1
        except Exception as e:
            console.print(f"  [!] Failed to create draft for {sender_email}: {str(e)[:50]}")
            drafts_failed += 1

    return drafts_created, drafts_failed

def _print_summary(drafts_created: int, drafts_failed: int):
    """Print draft creation summary"""
    console.print("\n" + "="*70)
    console.print("DRAFT CREATION SUMMARY")
    console.print("="*70)
    console.print(f"[+] Total drafts created: {drafts_created}")
    console.print(f"[+] Failed: {drafts_failed}")
    console.print("\n[+] Check your Gmail Drafts folder to review before sending!")
    console.print("="*70 + "\n")