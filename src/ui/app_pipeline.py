"""
Pipeline Page - Run all 4 agents sequentially

Author: Hadar Wayn
Date: December 2025
"""

import streamlit as st
import time
from .streamlit_components import render_agent1_form
from ..utils.paths import get_excel_file
from .agent_runner import _execute_agent1, _execute_agent2, _execute_agent3, _execute_agent4


def render_run_all_page():
    """Render Run All Agents page"""
    st.markdown("## 🚀 Run All Agents")
    st.warning("""
    **This will execute all 4 agents sequentially:**
    1. Agent 1: Email Extractor
    2. Agent 2: Repository Analyzer
    3. Agent 3: AI Feedback Generator
    4. Agent 4: Draft Creator
    """)
    params = render_agent1_form()
    if params:
        st.markdown("---")
        st.markdown("### 🔄 Pipeline Execution")
        st.info(f"""
        **Agent 1 Search Parameters:**
        - Subject: `{params.get('email_subject') or '(all subjects)'}`
        - Sender: `{params.get('sender_email') or '(all senders)'}`
        - Max Emails: `{params.get('max_emails', 20)}`
        """)
        agents = [
            ("Agent 1: Email Extractor", _execute_agent1, params),
            ("Agent 2: Repository Analyzer", _execute_agent2, {}),
            ("Agent 3: AI Feedback Generator", _execute_agent3, {}),
            ("Agent 4: Draft Creator", _execute_agent4, {})
        ]
        overall_progress = st.progress(0)
        results_summary = []
        for i, (name, func, agent_params) in enumerate(agents):
            st.markdown(f"#### {name}")
            result_placeholder = st.empty()
            with st.spinner(f"Running {name}..."):
                if i == 0:
                    success = func(agent_params)
                else:
                    success = func()
                if success:
                    if i == 0:
                        excel1 = get_excel_file('Excel1.xlsx')
                        if excel1.exists():
                            import openpyxl
                            wb = openpyxl.load_workbook(excel1)
                            ws = wb.active
                            row_count = ws.max_row - 1
                            ready_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[6] == 'Ready')
                            result_placeholder.success(f"✅ {name} completed! Found **{row_count}** emails (**{ready_count}** with GitHub URLs)")
                            results_summary.append(f"**Agent 1:** {row_count} emails ({ready_count} ready)")
                    elif i == 1:
                        excel2 = get_excel_file('Excel2.xlsx')
                        if excel2.exists():
                            import openpyxl
                            wb = openpyxl.load_workbook(excel2)
                            ws = wb.active
                            row_count = ws.max_row - 1
                            result_placeholder.success(f"✅ {name} completed! Analyzed **{row_count}** repositories")
                            results_summary.append(f"**Agent 2:** {row_count} repos analyzed")
                    elif i == 2:
                        excel3 = get_excel_file('Excel3.xlsx')
                        if excel3.exists():
                            import openpyxl
                            wb = openpyxl.load_workbook(excel3)
                            ws = wb.active
                            row_count = ws.max_row - 1
                            result_placeholder.success(f"✅ {name} completed! Generated **{row_count}** feedback messages")
                            results_summary.append(f"**Agent 3:** {row_count} feedback generated")
                    elif i == 3:
                        result_placeholder.success(f"✅ {name} completed! Gmail drafts created")
                        results_summary.append("**Agent 4:** Drafts created")
                else:
                    result_placeholder.error(f"❌ {name} failed!")
                    st.stop()
            overall_progress.progress((i + 1) / len(agents))
            time.sleep(0.5)
        st.success("🎉 All agents completed successfully!")
        st.balloons()
        st.markdown("### 📊 Pipeline Summary")
        for summary in results_summary:
            st.markdown(f"- {summary}")
        st.info("💡 Check the 'View Excel Files' page to see detailed results.")
