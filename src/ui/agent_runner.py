"""
Agent Runner - Invokes Claude Code agents via subprocess

This module handles the automated execution mode by invoking
Claude Code CLI to run agents defined in .claude/agents/

Author: Hadar Wayn
Date: December 2025
"""

import subprocess
from pathlib import Path
from typing import Optional

from .display import print_header, print_success, print_error, print_info, console
from ..utils.paths import get_project_root, get_claude_agents_dir


def run_agent(agent_name: str) -> bool:
    """
    Invoke Claude Code agent via CLI

    Args:
        agent_name: Name of agent (e.g., "Agent1_Email_Extractor")

    Returns:
        bool: True if successful, False otherwise

    Note:
        Agent 1 will prompt the user for search parameters when it runs:
        - Email Subject (optional)
        - Sender Email (optional)
        - Max Emails (1-100, default 10)
    """
    print_header(f"Running {agent_name}...")

    # Check if agent exists
    agent_dir = get_claude_agents_dir() / agent_name
    skill_file = agent_dir / "SKILL.md"

    if not skill_file.exists():
        print_error(f"Agent not found: {skill_file}")
        return False

    print_info(f"Agent location: {skill_file}")

    # Collect parameters for Agent 1
    agent_params = {}
    if agent_name == "Agent1_Email_Extractor":
        agent_params = _prompt_agent1_parameters()

        # Execute Agent 1 directly
        console.print("\n[bold green][*] Executing Agent 1...[/bold green]\n")
        success = _execute_agent1(agent_params)

        if success:
            print_success("Agent 1 completed successfully")
        else:
            print_error("Agent 1 execution failed")

        console.input("\n[dim]Press Enter to continue...[/dim]")
        return success

    # Execute Agent 2 directly
    if agent_name == "Agent2_Repository_Analyzer":
        console.print("\n[bold green][*] Executing Agent 2...[/bold green]\n")
        success = _execute_agent2()

        if success:
            print_success("Agent 2 completed successfully")
        else:
            print_error("Agent 2 execution failed")

        console.input("\n[dim]Press Enter to continue...[/dim]")
        return success

    # Execute Agent 3 directly
    if agent_name == "Agent3_LLM_Feedback":
        console.print("\n[bold green][*] Executing Agent 3...[/bold green]\n")
        success = _execute_agent3()

        if success:
            print_success("Agent 3 completed successfully")
        else:
            print_error("Agent 3 execution failed")

        console.input("\n[dim]Press Enter to continue...[/dim]")
        return success

    # Execute Agent 4 directly
    if agent_name == "Agent4_Draft_Creator":
        console.print("\n[bold green][*] Executing Agent 4...[/bold green]\n")
        success = _execute_agent4()

        if success:
            print_success("Agent 4 completed successfully")
        else:
            print_error("Agent 4 execution failed")

        console.input("\n[dim]Press Enter to continue...[/dim]")
        return success

    # For other agents, display manual instructions
    console.print("\n[bold yellow][*] Manual Execution Required:[/bold yellow]")
    console.print(f"\n  To run this agent manually, execute:")
    console.print(f"  [cyan]claude[/cyan]")
    console.print(f"  [cyan]/agents[/cyan]")
    console.print(f"  Select: [green]{agent_name}[/green]\n")

    console.print("[yellow][!] Automated execution not yet implemented for this agent.[/yellow]")
    console.print("[yellow]   This will be added in a future update.[/yellow]\n")

    # Simulated success for now
    console.input("\n[dim]Press Enter to continue...[/dim]")
    return True


def _prompt_agent1_parameters() -> dict:
    """
    Prompt user for Agent 1 search parameters

    Returns:
        dict: Dictionary containing email_subject, sender_email, and max_emails
    """
    console.print("\n[bold cyan]=====================================================================[/bold cyan]")
    console.print("[bold cyan]              AGENT 1: EMAIL EXTRACTION PARAMETERS                [/bold cyan]")
    console.print("[bold cyan]=====================================================================[/bold cyan]\n")

    # Email subject (optional)
    email_subject = console.input("[bold]Email Subject[/bold] (optional, press Enter to skip): ").strip()
    if not email_subject:
        email_subject = None
        console.print("   -> Will search all email subjects\n")
    else:
        console.print(f"   -> Will search for: [green]\"{email_subject}\"[/green]\n")

    # Sender email (optional)
    sender_email = console.input("[bold]Sender Email[/bold] (optional, press Enter to skip): ").strip()
    if not sender_email:
        sender_email = None
        console.print("   -> Will search all senders\n")
    else:
        # Basic email validation
        if '@' not in sender_email or '.' not in sender_email.split('@')[-1]:
            console.print("   [yellow][!] Invalid email format, will search all senders[/yellow]\n")
            sender_email = None
        else:
            console.print(f"   -> Will filter by sender: [green]{sender_email}[/green]\n")

    # Max emails (required, default 10)
    while True:
        max_emails_input = console.input("[bold]Max Emails (1-100)[/bold] [dim][Default: 10][/dim]: ").strip()
        if not max_emails_input:
            max_emails = 10
            console.print("   -> Using default: [green]10 emails[/green]\n")
            break
        else:
            try:
                max_emails = int(max_emails_input)
                if 1 <= max_emails <= 100:
                    console.print(f"   -> Will process up to [green]{max_emails} emails[/green]\n")
                    break
                else:
                    console.print("   [red][!] Please enter a number between 1 and 100[/red]")
            except ValueError:
                console.print("   [red][!] Please enter a valid number[/red]")

    # Build Gmail search query
    query_parts = []
    if email_subject:
        # Gmail search WITHOUT quotes works best - it finds emails containing all words
        # Python post-filter ensures exact substring match (case-insensitive)
        # Use first 3-4 words for broad Gmail search, post-filter ensures accuracy
        words = email_subject.split()

        if len(words) <= 3:
            # Short search (1-3 words): use all words
            query_parts.append(f'subject:{email_subject}')
        else:
            # Longer search (4+ words): use first 3 words for broad Gmail match
            # Post-filter will ensure full phrase is in subject
            key_words = ' '.join(words[:3])
            query_parts.append(f'subject:{key_words}')

    if sender_email:
        query_parts.append(f'from:{sender_email}')

    # Combine query parts (search all emails if no filters)
    query = ' '.join(query_parts) if query_parts else ''

    console.print("[bold cyan]=====================================================================[/bold cyan]")
    console.print(f"[bold]Gmail Search Query:[/bold] [cyan]{query}[/cyan]")
    console.print(f"[bold]Max Results:[/bold] [cyan]{max_emails}[/cyan]")
    console.print("[bold cyan]=====================================================================[/bold cyan]\n")

    return {
        'email_subject': email_subject,
        'sender_email': sender_email,
        'max_emails': max_emails,
        'query': query
    }


def _execute_agent1(params: dict) -> bool:
    """
    Execute Agent 1 email extraction logic

    Args:
        params: Dictionary with email_subject, sender_email, max_emails, query

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        import openpyxl
        import re
        import hashlib
        from email.utils import parsedate_to_datetime
        import base64

        # Gmail API scopes
        SCOPES = [
            'https://www.googleapis.com/auth/gmail.readonly',
            'https://www.googleapis.com/auth/gmail.compose'
        ]

        # Extract parameters
        email_subject = params.get('email_subject')
        sender_email = params.get('sender_email')
        max_emails = params.get('max_emails', 10)
        query = params.get('query', '')

        console.print(f"[*] Email Subject: {email_subject or '(all subjects)'}")
        console.print(f"[*] Sender Email: {sender_email or '(all senders)'}")
        console.print(f"[*] Max Emails: {max_emails}")
        console.print(f"[*] Gmail Query: {query}\n")

        # Gmail Authentication
        console.print("[*] Authenticating with Gmail API...")
        creds = None
        token_path = Path('Secrets/token.json')
        credentials_path = Path('Secrets/credentials.json')

        if token_path.exists():
            creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCOPES)
                creds = flow.run_local_server(port=0)

            with open(token_path, 'w') as token:
                token.write(creds.to_json())

        console.print("[+] Authentication successful\n")

        # Create Gmail service
        service = build('gmail', 'v1', credentials=creds)

        # Search emails
        console.print("[*] Searching emails...")
        results = service.users().messages().list(
            userId='me',
            q=query,
            maxResults=max_emails
        ).execute()

        messages = results.get('messages', [])

        if not messages:
            console.print("[!] No emails found matching the criteria\n")
            console.print("[*] Creating empty Excel file with headers...\n")
            submissions = []
        else:
            console.print(f"[+] Found {len(messages)} emails\n")

            # Extract Data
            submissions = []
            for i, message in enumerate(messages, 1):
                try:
                    console.print(f"Processing email {i}/{len(messages)}...")

                    msg = service.users().messages().get(userId='me', id=message['id'], format='full').execute()

                    headers = msg['payload']['headers']
                    sender = next((h['value'] for h in headers if h['name'] == 'From'), '')
                    subject = next((h['value'] for h in headers if h['name'] == 'Subject'), '')
                    date_str = next((h['value'] for h in headers if h['name'] == 'Date'), '')

                    # Post-filter: if email_subject was specified, verify it's a substring (case-insensitive)
                    if email_subject and email_subject.lower() not in subject.lower():
                        console.print(f"  [-] Subject doesn't contain '{email_subject}', skipping")
                        continue

                    email_match = re.search(r'<(.+?)>', sender)
                    sender_email_extracted = email_match.group(1) if email_match else sender

                    received_time = parsedate_to_datetime(date_str).isoformat()
                    thread_id = msg['threadId']

                    def get_body(payload):
                        if 'body' in payload and 'data' in payload['body']:
                            return base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8', errors='ignore')
                        if 'parts' in payload:
                            for part in payload['parts']:
                                body = get_body(part)
                                if body:
                                    return body
                        return ''

                    body = get_body(msg['payload'])

                    github_pattern = r'https://github\.com/([a-zA-Z0-9_-]+)/([a-zA-Z0-9_.-]+)(\.git)?'
                    github_match = re.search(github_pattern, body)

                    email_id = hashlib.sha256(f"{sender_email_extracted}|{subject}|{received_time}".encode()).hexdigest()

                    if github_match:
                        status = "Ready"
                        github_url = github_match.group(0)
                        console.print(f"  [+] Found GitHub URL: {github_url}")
                    else:
                        status = "Missing: github_url"
                        github_url = ""
                        console.print(f"  [!] No GitHub URL found")

                    submissions.append({
                        'email_id': email_id,
                        'received_time': received_time,
                        'email_subject': subject,
                        'sender_email': sender_email_extracted,
                        'github_url': github_url,
                        'thread_id': thread_id,
                        'status': status
                    })

                except Exception as e:
                    console.print(f"  [-] Error processing email: {e}")
                    submissions.append({
                        'email_id': f"error_{i}",
                        'received_time': '',
                        'email_subject': '',
                        'sender_email': '',
                        'github_url': '',
                        'thread_id': '',
                        'status': f"Error: {str(e)}"
                    })

        # Create Excel1.xlsx
        console.print("\n[*] Creating Excel1.xlsx...")

        from ..utils.paths import get_excel_dir
        excel_dir = get_excel_dir()
        excel_dir.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Email Submissions"

        headers = ["email_id", "received_time", "email_subject", "sender_email",
                   "github_url", "thread_id", "status"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for submission in submissions:
            ws.append([
                submission['email_id'],
                submission['received_time'],
                submission['email_subject'],
                submission['sender_email'],
                submission['github_url'],
                submission['thread_id'],
                submission['status']
            ])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output_path = excel_dir / 'Excel1.xlsx'
        wb.save(output_path)

        # Output Summary
        ready_count = sum(1 for s in submissions if s['status'] == "Ready")
        missing_count = sum(1 for s in submissions if "Missing" in s['status'])
        error_count = sum(1 for s in submissions if "Error" in s['status'])

        console.print("\n" + "="*70)
        console.print("EXTRACTION SUMMARY")
        console.print("="*70)
        console.print(f"[+] Processed: {len(submissions)} emails")
        console.print(f"[+] Excel1.xlsx created: {output_path}")
        console.print(f"\nStatus breakdown:")
        console.print(f"   - Ready (with GitHub URL): {ready_count}")
        console.print(f"   - Missing GitHub URL: {missing_count}")
        console.print(f"   - Errors: {error_count}")
        console.print("="*70 + "\n")

        return True

    except Exception as e:
        console.print(f"\n[red][!] Agent 1 execution failed: {e}[/red]\n")
        import traceback
        console.print(f"[dim]{traceback.format_exc()}[/dim]")
        return False


def _execute_agent2() -> bool:
    """
    Execute Agent 2 repository analysis logic

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        import openpyxl
        import git
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import shutil

        from ..utils.paths import get_excel_file, get_excel_dir

        # Check if Excel1.xlsx exists
        excel1_path = get_excel_file('Excel1.xlsx')
        if not excel1_path.exists():
            console.print("[red][!] Excel1.xlsx not found. Please run Agent 1 first.[/red]\n")
            return False

        console.print(f"[*] Reading Excel1.xlsx: {excel1_path}\n")

        # Load Excel1
        wb = openpyxl.load_workbook(excel1_path)
        ws = wb.active

        # Extract rows with status = "Ready"
        ready_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            email_id, received_time, subject, sender, github_url, thread_id, status = row
            if status == "Ready":
                ready_rows.append({
                    'email_id': email_id,
                    'github_url': github_url
                })

        console.print(f"[+] Found {len(ready_rows)} repositories to analyze\n")

        if len(ready_rows) == 0:
            console.print("[!] No repositories to analyze. Creating empty Excel2.xlsx...\n")
            results = []
        else:
            # Define analysis function
            def analyze_repository(email_id, github_url):
                """Analyze a single repository"""
                try:
                    # Create unique directory for this repo
                    repo_dir = Path(f"temp/repos/{email_id[:8]}")

                    # Remove directory if it exists
                    if repo_dir.exists():
                        shutil.rmtree(repo_dir)

                    # Clone repository (depth=1 for speed)
                    console.print(f"[*] Cloning {github_url[:60]}...")
                    git.Repo.clone_from(github_url, repo_dir, depth=1)

                    # Find all Python files
                    python_files = list(repo_dir.rglob("*.py"))

                    # Count lines
                    total_lines = 0
                    compliant_lines = 0
                    file_count = 0

                    for py_file in python_files:
                        try:
                            with open(py_file, 'r', encoding='utf-8') as f:
                                lines = len(f.readlines())
                                total_lines += lines

                                # Compliant if <= 150 lines
                                if lines <= 150:
                                    compliant_lines += lines

                                file_count += 1
                        except Exception:
                            continue

                    # Calculate grade
                    if total_lines > 0:
                        grade = round(100 * (compliant_lines / total_lines), 2)
                    else:
                        grade = 0

                    console.print(f"  [+] {email_id[:8]} - Files: {file_count}, Grade: {grade}%")

                    return {
                        'email_id': email_id,
                        'github_url': github_url,
                        'total_files': file_count,
                        'total_lines': total_lines,
                        'compliant_lines': compliant_lines,
                        'grade': grade,
                        'status': 'Ready'
                    }

                except git.GitCommandError as e:
                    console.print(f"  [!] {email_id[:8]} - Clone failed")
                    return {
                        'email_id': email_id,
                        'github_url': github_url,
                        'total_files': 0,
                        'total_lines': 0,
                        'compliant_lines': 0,
                        'grade': 0,
                        'status': 'Failed: clone'
                    }
                except Exception as e:
                    console.print(f"  [!] {email_id[:8]} - Error: {str(e)[:50]}")
                    return {
                        'email_id': email_id,
                        'github_url': github_url,
                        'total_files': 0,
                        'total_lines': 0,
                        'compliant_lines': 0,
                        'grade': 0,
                        'status': f'Failed: {str(e)[:50]}'
                    }

            # Process repositories in parallel (5 workers)
            results = []
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {
                    executor.submit(analyze_repository, row['email_id'], row['github_url']): row
                    for row in ready_rows
                }

                for future in as_completed(futures):
                    result = future.result()
                    results.append(result)

        # Create Excel2.xlsx
        console.print("\n[*] Creating Excel2.xlsx...")

        excel_dir = get_excel_dir()
        excel_dir.mkdir(parents=True, exist_ok=True)

        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Repository Analysis"

        # Headers
        headers = [
            "email_id",
            "github_url",
            "total_files",
            "total_lines",
            "compliant_lines",
            "grade",
            "status"
        ]
        ws2.append(headers)

        # Format headers
        for cell in ws2[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Add rows
        for result in results:
            ws2.append([
                result['email_id'],
                result['github_url'],
                result['total_files'],
                result['total_lines'],
                result['compliant_lines'],
                result['grade'],
                result['status']
            ])

        # Auto-adjust column widths
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width

        # Save
        output_path = excel_dir / 'Excel2.xlsx'
        wb2.save(output_path)

        # Output Summary
        successful_count = sum(1 for r in results if r['status'] == 'Ready')
        failed_count = sum(1 for r in results if 'Failed' in r['status'])
        avg_grade = sum(r['grade'] for r in results) / len(results) if results else 0

        console.print("\n" + "="*70)
        console.print("ANALYSIS SUMMARY")
        console.print("="*70)
        console.print(f"[+] Total repositories: {len(results)}")
        console.print(f"[+] Successful: {successful_count}")
        console.print(f"[+] Failed: {failed_count}")
        console.print(f"[+] Average grade: {avg_grade:.2f}%")
        console.print(f"[+] Excel2.xlsx created: {output_path}")
        console.print("="*70 + "\n")

        return True

    except Exception as e:
        console.print(f"\n[red][!] Agent 2 execution failed: {e}[/red]\n")
        import traceback
        console.print(f"[dim]{traceback.format_exc()}[/dim]")
        return False


def _execute_agent3() -> bool:
    """
    Execute Agent 3 LLM feedback generation logic

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        import openpyxl
        import google.generativeai as genai
        import os
        import time
        from dotenv import load_dotenv

        from ..utils.paths import get_excel_file, get_excel_dir

        # Load environment variables
        load_dotenv()

        # Check if Excel2.xlsx exists
        excel2_path = get_excel_file('Excel2.xlsx')
        if not excel2_path.exists():
            console.print("[red][!] Excel2.xlsx not found. Please run Agent 2 first.[/red]\n")
            return False

        # Check Gemini API key
        gemini_api_key = os.getenv('GEMINI_API_KEY')
        if not gemini_api_key:
            console.print("[red][!] GEMINI_API_KEY not found in environment variables.[/red]\n")
            console.print("[yellow]Please set it in your .env file or environment.[/yellow]\n")
            return False

        console.print(f"[*] Reading Excel2.xlsx: {excel2_path}\n")

        # Configure Gemini
        genai.configure(api_key=gemini_api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')

        # Load Excel2
        wb = openpyxl.load_workbook(excel2_path)
        ws = wb.active

        # Extract rows with status = "Ready"
        ready_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            email_id, github_url, total_files, total_lines, compliant_lines, grade, status = row
            if status == "Ready":
                ready_rows.append({
                    'email_id': email_id,
                    'github_url': github_url,
                    'grade': grade
                })

        console.print(f"[+] Found {len(ready_rows)} submissions to generate feedback for\n")

        if len(ready_rows) == 0:
            console.print("[!] No submissions to process. Creating empty Excel3.xlsx...\n")
            results = []
        else:
            # Define helper functions
            def get_grade_category(grade):
                """Determine grade category"""
                if 90 <= grade <= 100:
                    return 'excellent', 'trump'
                elif 70 <= grade < 90:
                    return 'good', 'hason'
                elif 55 <= grade < 70:
                    return 'pass', 'lee'
                else:
                    return 'needs_work', 'amsalem'

            def load_persona_prompt(persona_name):
                """Load persona skill file and extract prompt template"""
                skill_path = Path(f'.claude/agents/Agent3_LLM_Feedback/personas/{persona_name}_skill.md')

                with open(skill_path, 'r', encoding='utf-8') as f:
                    skill_content = f.read()

                return skill_content

            def generate_feedback_with_retry(persona_name, grade, github_url, max_retries=3):
                """Generate feedback with exponential backoff retry"""

                # Load persona prompt
                persona_prompt = load_persona_prompt(persona_name)

                # Create final prompt
                prompt = f"""
{persona_prompt}

Assignment Details:
- Repository: {github_url}
- Grade: {grade}/100
- Grade based on code structure compliance (files under 150 lines)

Generate a personalized feedback message (2-3 sentences) in the persona's style.
Focus on encouragement and specific advice based on the grade.
"""

                # Retry loop
                for attempt in range(max_retries):
                    try:
                        # Generate response
                        response = model.generate_content(prompt)
                        text = response.text.strip()

                        if text:
                            return text, attempt + 1

                    except Exception as e:
                        console.print(f"  [!] Attempt {attempt + 1} failed: {str(e)[:50]}")

                        if attempt < max_retries - 1:
                            # Exponential backoff: 1s, 2s, 4s
                            wait_time = 2 ** attempt
                            time.sleep(wait_time)
                        else:
                            # Final attempt failed
                            return "", max_retries

                return "", max_retries

            # Process all submissions
            results = []

            for row in ready_rows:
                email_id = row['email_id']
                grade = row['grade']
                github_url = row['github_url']

                # Determine persona
                category, persona = get_grade_category(grade)

                console.print(f"[*] Processing {email_id[:8]} - Grade: {grade}% - Persona: {persona}")

                # Generate feedback
                response, attempts = generate_feedback_with_retry(persona, grade, github_url)

                # Determine status
                if response:
                    status = "Ready"
                    console.print(f"  [+] Generated feedback ({attempts} attempt(s))")
                else:
                    status = "Missing: reply"
                    console.print(f"  [!] Failed to generate feedback")

                results.append({
                    'email_id': email_id,
                    'grade': grade,
                    'grade_category': category,
                    'persona': persona,
                    'response': response,
                    'api_attempts': attempts,
                    'status': status
                })

                # Delay between API calls
                time.sleep(2)

        # Create Excel3.xlsx
        console.print("\n[*] Creating Excel3.xlsx...")

        excel_dir = get_excel_dir()
        excel_dir.mkdir(parents=True, exist_ok=True)

        wb3 = openpyxl.Workbook()
        ws3 = wb3.active
        ws3.title = "LLM Feedback"

        # Headers
        headers = [
            "email_id",
            "grade",
            "grade_category",
            "persona",
            "response",
            "api_attempts",
            "status"
        ]
        ws3.append(headers)

        # Format headers
        for cell in ws3[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Add rows
        for result in results:
            ws3.append([
                result['email_id'],
                result['grade'],
                result['grade_category'],
                result['persona'],
                result['response'],
                result['api_attempts'],
                result['status']
            ])

        # Auto-adjust column widths
        for column in ws3.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws3.column_dimensions[column_letter].width = adjusted_width

        # Save
        output_path = excel_dir / 'Excel3.xlsx'
        wb3.save(output_path)

        # Output Summary
        successful_count = sum(1 for r in results if r['status'] == 'Ready')
        failed_count = sum(1 for r in results if 'Missing' in r['status'])

        console.print("\n" + "="*70)
        console.print("FEEDBACK GENERATION SUMMARY")
        console.print("="*70)
        console.print(f"[+] Total submissions: {len(results)}")
        console.print(f"[+] Successful: {successful_count}")
        console.print(f"[+] Failed: {failed_count}")

        if results:
            console.print(f"\nPersonas used:")
            for persona in ['trump', 'hason', 'lee', 'amsalem']:
                count = sum(1 for r in results if r['persona'] == persona)
                if count > 0:
                    console.print(f"   - {persona}: {count}")

        console.print(f"\n[+] Excel3.xlsx created: {output_path}")
        console.print("="*70 + "\n")

        return True

    except Exception as e:
        console.print(f"\n[red][!] Agent 3 execution failed: {e}[/red]\n")
        import traceback
        console.print(f"[dim]{traceback.format_exc()}[/dim]")
        return False


def _execute_agent4() -> bool:
    """
    Execute Agent 4 draft email creation logic

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        import openpyxl
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        from google.auth.transport.requests import Request
        from google_auth_oauthlib.flow import InstalledAppFlow
        from email.mime.text import MIMEText
        import base64

        from ..utils.paths import get_excel_file

        # Check if Excel3.xlsx exists
        excel3_path = get_excel_file('Excel3.xlsx')
        if not excel3_path.exists():
            console.print("[red][!] Excel3.xlsx not found. Please run Agent 3 first.[/red]\n")
            return False

        # Check if Excel1.xlsx exists
        excel1_path = get_excel_file('Excel1.xlsx')
        if not excel1_path.exists():
            console.print("[red][!] Excel1.xlsx not found. Please run Agent 1 first.[/red]\n")
            return False

        console.print(f"[*] Reading Excel3.xlsx: {excel3_path}")
        console.print(f"[*] Reading Excel1.xlsx: {excel1_path}\n")

        # Load Excel3 (feedback)
        wb3 = openpyxl.load_workbook(excel3_path)
        ws3 = wb3.active

        feedback_data = {}
        for row in ws3.iter_rows(min_row=2, values_only=True):
            email_id, grade, category, persona, response, attempts, status = row
            if status == "Ready" and response:
                feedback_data[email_id] = {
                    'grade': grade,
                    'response': response
                }

        console.print(f"[+] Found {len(feedback_data)} submissions with feedback\n")

        if len(feedback_data) == 0:
            console.print("[!] No submissions with feedback to process.\n")
            return True

        # Load Excel1 (email metadata)
        wb1 = openpyxl.load_workbook(excel1_path)
        ws1 = wb1.active

        email_metadata = {}
        for row in ws1.iter_rows(min_row=2, values_only=True):
            email_id, received_time, subject, sender, github_url, thread_id, status = row
            email_metadata[email_id] = {
                'sender': sender,
                'subject': subject,
                'thread_id': thread_id,
                'github_url': github_url
            }

        # Load student mapping (optional)
        student_names = {}
        mapping_path = Path('data/students_mapping.xlsx')
        if mapping_path.exists():
            try:
                wb_mapping = openpyxl.load_workbook(mapping_path)
                ws_mapping = wb_mapping.active

                for row in ws_mapping.iter_rows(min_row=2, values_only=True):
                    email, name = row
                    if email and name:
                        student_names[email.lower().strip()] = name

                console.print(f"[+] Loaded {len(student_names)} student name mappings\n")
            except Exception as e:
                console.print(f"[!] Could not load student mappings: {e}\n")

        # Gmail Authentication
        console.print("[*] Authenticating with Gmail API...")
        SCOPES = [
            'https://www.googleapis.com/auth/gmail.readonly',
            'https://www.googleapis.com/auth/gmail.compose'
        ]

        creds = None
        token_path = Path('Secrets/token.json')
        credentials_path = Path('Secrets/credentials.json')

        if token_path.exists():
            creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCOPES)
                creds = flow.run_local_server(port=0)

            with open(token_path, 'w') as token:
                token.write(creds.to_json())

        console.print("[+] Authentication successful\n")

        # Create Gmail service
        service = build('gmail', 'v1', credentials=creds)

        # Create drafts
        console.print("[*] Creating Gmail drafts...\n")
        drafts_created = 0
        drafts_failed = 0

        for email_id, feedback in feedback_data.items():
            # Get email metadata
            metadata = email_metadata.get(email_id)
            if not metadata:
                console.print(f"  [!] No metadata for {email_id[:8]}, skipping")
                drafts_failed += 1
                continue

            sender_email = metadata['sender']
            original_subject = metadata['subject']
            thread_id = metadata['thread_id']
            github_url = metadata['github_url']
            grade = feedback['grade']
            response_text = feedback['response']

            # Get student name (or use email prefix)
            student_name = student_names.get(sender_email.lower(), sender_email.split('@')[0])

            # Compose email
            email_body = f"""Hi {student_name}!

{response_text}

Your repository reviewed: {github_url}
Your grade: {grade}/100

Best regards,
Instructor
"""

            # Create MIME message
            message = MIMEText(email_body)
            message['to'] = sender_email
            message['subject'] = f"Re: {original_subject}"

            # Encode message
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')

            # Create draft as reply (using thread_id)
            try:
                draft = {
                    'message': {
                        'raw': raw_message,
                        'threadId': thread_id
                    }
                }

                created_draft = service.users().drafts().create(
                    userId='me',
                    body=draft
                ).execute()

                console.print(f"  [+] Draft created for {student_name} ({sender_email})")
                drafts_created += 1

            except Exception as e:
                console.print(f"  [!] Failed to create draft for {sender_email}: {str(e)[:50]}")
                drafts_failed += 1

        # Output Summary
        console.print("\n" + "="*70)
        console.print("DRAFT CREATION SUMMARY")
        console.print("="*70)
        console.print(f"[+] Total drafts created: {drafts_created}")
        console.print(f"[+] Failed: {drafts_failed}")
        console.print("\n[+] Check your Gmail Drafts folder to review before sending!")
        console.print("="*70 + "\n")

        return True

    except Exception as e:
        console.print(f"\n[red][!] Agent 4 execution failed: {e}[/red]\n")
        import traceback
        console.print(f"[dim]{traceback.format_exc()}[/dim]")
        return False


def run_all_agents() -> bool:
    """
    Run all 4 agents sequentially

    Returns:
        bool: True if all successful, False otherwise

    Note:
        Agent 1 will prompt for search parameters at the start
    """
    print_header("Running All Agents Sequentially")

    agents = [
        "Agent1_Email_Extractor",  # Prompts user for parameters
        "Agent2_Repository_Analyzer",
        "Agent3_LLM_Feedback",
        "Agent4_Draft_Creator"
    ]

    for agent_name in agents:
        success = run_agent(agent_name)

        if not success:
            print_error(f"Agent {agent_name} failed. Stopping pipeline.")
            return False

        # Verify output exists
        if not verify_agent_output(agent_name):
            print_error(f"Agent {agent_name} did not produce expected output. Stopping.")
            return False

    print_success("🎉 All agents completed successfully!")
    return True


def verify_agent_output(agent_name: str) -> bool:
    """
    Verify agent produced expected output

    Args:
        agent_name: Name of agent

    Returns:
        bool: True if output exists, False otherwise
    """
    from ..utils.paths import get_excel_file

    expected_files = {
        "Agent1_Email_Extractor": "Excel1.xlsx",
        "Agent2_Repository_Analyzer": "Excel2.xlsx",
        "Agent3_LLM_Feedback": "Excel3.xlsx",
        "Agent4_Draft_Creator": None  # Creates Gmail drafts, no file
    }

    expected_file = expected_files.get(agent_name)

    if expected_file is None:
        # Agent 4 doesn't create a file
        return True

    file_path = get_excel_file(expected_file)
    exists = file_path.exists()

    if exists:
        print_info(f"Verified: {expected_file} exists")
    else:
        print_error(f"Missing: {expected_file}")

    return exists
