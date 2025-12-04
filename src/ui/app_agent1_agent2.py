"""
Agent 1 & 2 Pages - Email Extractor and Repository Analyzer

Author: Hadar Wayn
Date: December 2025
"""

import streamlit as st
import time
from .streamlit_components import render_agent1_form, render_excel_viewer
from ..utils.paths import get_excel_file
from .agent_runner import _execute_agent1, _execute_agent2


def render_agent1_page():
    """Render Agent 1 execution page with form"""
    st.markdown("## 📧 Agent 1: Email Extractor")
    st.info("""
    **What this agent does:**
    - Connects to Gmail API
    - Searches for emails matching your criteria
    - Extracts GitHub URLs from email bodies
    - Creates Excel1.xlsx with extracted data
    """)
    params = render_agent1_form()
    if params:
        st.markdown("---")
        st.markdown("### 🔄 Execution")
        with st.spinner("Executing Agent 1..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            status_text.text("Authenticating with Gmail...")
            progress_bar.progress(20)
            time.sleep(0.5)
            status_text.text("Searching emails...")
            progress_bar.progress(40)
            success = _execute_agent1(params)
            progress_bar.progress(100)
            if success:
                excel_path = get_excel_file('Excel1.xlsx')
                if excel_path.exists():
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    row_count = ws.max_row - 1
                    ready_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[6] == 'Ready')
                    st.success(f"✅ Agent 1 completed! Found **{row_count}** emails (**{ready_count}** with GitHub URLs)")
                    render_excel_viewer(excel_path, "Excel1.xlsx - Email Extraction Results")
                    if ready_count > 0:
                        st.info("💡 **Next Step:** Run Agent 2 to analyze the repositories")
                        if st.button("▶️ Go to Agent 2", use_container_width=True):
                            st.session_state.current_page = 'agent2'
                            st.rerun()
                    else:
                        st.warning("⚠️ No emails with GitHub URLs found. Try adjusting your search parameters.")
            else:
                st.error("❌ Agent 1 failed. Check the logs for details.")


def render_agent2_page():
    """Render Agent 2 execution page"""
    st.markdown("## 🔬 Agent 2: Repository Analyzer")
    st.info("""
    **What this agent does:**
    - Reads Excel1.xlsx for GitHub URLs
    - Clones repositories (5 concurrent workers)
    - Analyzes Python code quality
    - Calculates grades based on line limits
    - Creates Excel2.xlsx with analysis results
    """)
    excel1 = get_excel_file('Excel1.xlsx')
    if not excel1.exists():
        st.warning("⚠️ Excel1.xlsx not found. Please run Agent 1 first.")
        if st.button("▶️ Go to Agent 1", use_container_width=True):
            st.session_state.current_page = 'agent1'
            st.rerun()
        return
    if st.button("🚀 Run Agent 2", use_container_width=True, type="primary"):
        st.markdown("---")
        st.markdown("### 🔄 Execution")
        with st.spinner("Executing Agent 2..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            status_text.text("Reading Excel1.xlsx...")
            progress_bar.progress(10)
            time.sleep(0.5)
            status_text.text("Cloning repositories (5 workers)...")
            progress_bar.progress(30)
            success = _execute_agent2()
            progress_bar.progress(100)
            if success:
                excel_path = get_excel_file('Excel2.xlsx')
                if excel_path.exists():
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    row_count = ws.max_row - 1
                    st.success(f"✅ Agent 2 completed! Analyzed **{row_count}** repositories")
                    render_excel_viewer(excel_path, "Excel2.xlsx - Repository Analysis")
                    if row_count > 0:
                        st.info("💡 **Next Step:** Run Agent 3 to generate AI feedback")
                        if st.button("▶️ Go to Agent 3", use_container_width=True):
                            st.session_state.current_page = 'agent3'
                            st.rerun()
            else:
                st.error("❌ Agent 2 failed. Check the logs for details.")
