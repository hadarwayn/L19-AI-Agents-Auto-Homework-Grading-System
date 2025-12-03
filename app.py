"""
L19 AI Agents Auto Homework Grading System - Streamlit Web UI
Modern, interactive web interface with clickable buttons and form fields
"""

import streamlit as st
import sys
from pathlib import Path
import subprocess
import time

# Add src to path
sys.path.insert(0, str(Path(__file__).parent))

from src.ui.streamlit_components import (
    render_header,
    render_agent_card,
    render_agent1_form,
    render_progress_tracker,
    render_excel_viewer,
    render_system_status,
    render_status_cards
)
from src.utils.paths import get_excel_file, get_temp_dir, get_results_dir, clean_temp_repos, get_excel_dir
from src.ui.agent_runner import (
    _execute_agent1,
    _execute_agent2,
    _execute_agent3,
    _execute_agent4
)


def main():
    """Main Streamlit application"""
    render_header()

    # Initialize session state
    if 'current_page' not in st.session_state:
        st.session_state.current_page = 'dashboard'
    if 'agent1_params' not in st.session_state:
        st.session_state.agent1_params = None

    # Sidebar navigation
    with st.sidebar:
        st.markdown("## 🎯 Navigation")

        if st.button("🏠 Dashboard", use_container_width=True):
            st.session_state.current_page = 'dashboard'
            st.rerun()

        if st.button("📧 Agent 1: Email Extractor", use_container_width=True):
            st.session_state.current_page = 'agent1'
            st.rerun()

        if st.button("🔬 Agent 2: Repository Analyzer", use_container_width=True):
            st.session_state.current_page = 'agent2'
            st.rerun()

        if st.button("🤖 Agent 3: AI Feedback Generator", use_container_width=True):
            st.session_state.current_page = 'agent3'
            st.rerun()

        if st.button("✉️ Agent 4: Draft Creator", use_container_width=True):
            st.session_state.current_page = 'agent4'
            st.rerun()

        if st.button("🚀 Run All Agents", use_container_width=True, type="primary"):
            st.session_state.current_page = 'run_all'
            st.rerun()

        st.markdown("---")

        if st.button("📊 View Excel Files", use_container_width=True):
            st.session_state.current_page = 'view_data'
            st.rerun()

        if st.button("ℹ️ System Status", use_container_width=True):
            st.session_state.current_page = 'status'
            st.rerun()

        st.markdown("---")

        if st.button("🗑️ Reset System", use_container_width=True):
            st.session_state.current_page = 'reset'
            st.rerun()

    # Main content area
    if st.session_state.current_page == 'dashboard':
        render_dashboard()
    elif st.session_state.current_page == 'agent1':
        render_agent1_page()
    elif st.session_state.current_page == 'agent2':
        render_agent2_page()
    elif st.session_state.current_page == 'agent3':
        render_agent3_page()
    elif st.session_state.current_page == 'agent4':
        render_agent4_page()
    elif st.session_state.current_page == 'run_all':
        render_run_all_page()
    elif st.session_state.current_page == 'view_data':
        render_view_data_page()
    elif st.session_state.current_page == 'status':
        render_status_page()
    elif st.session_state.current_page == 'reset':
        render_reset_page()


def render_dashboard():
    """Render the main dashboard with agent cards"""
    st.markdown("## 🏠 Dashboard")

    # System status cards
    status_data = render_system_status()
    render_status_cards(status_data)

    st.markdown("---")

    # Agent cards
    st.markdown("## 🤖 Agents")

    col1, col2 = st.columns(2)

    with col1:
        with st.container():
            st.markdown("""
            <div class="agent-card">
                <h3>📧 Agent 1: Email Extractor</h3>
                <p>Extract homework submissions from Gmail with flexible search parameters</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("▶️ Run Agent 1", key="dash_agent1", use_container_width=True):
                st.session_state.current_page = 'agent1'
                st.rerun()

        with st.container():
            st.markdown("""
            <div class="agent-card">
                <h3>🤖 Agent 3: AI Feedback</h3>
                <p>Generate personalized feedback using 4 celebrity personas</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("▶️ Run Agent 3", key="dash_agent3", use_container_width=True):
                st.session_state.current_page = 'agent3'
                st.rerun()

    with col2:
        with st.container():
            st.markdown("""
            <div class="agent-card">
                <h3>🔬 Agent 2: Repository Analyzer</h3>
                <p>Clone and analyze GitHub repositories with multi-threading</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("▶️ Run Agent 2", key="dash_agent2", use_container_width=True):
                st.session_state.current_page = 'agent2'
                st.rerun()

        with st.container():
            st.markdown("""
            <div class="agent-card">
                <h3>✉️ Agent 4: Draft Creator</h3>
                <p>Create Gmail draft replies with AI-generated feedback</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("▶️ Run Agent 4", key="dash_agent4", use_container_width=True):
                st.session_state.current_page = 'agent4'
                st.rerun()

    st.markdown("---")

    # Quick actions
    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("🚀 Run All Agents", key="dash_run_all", use_container_width=True, type="primary"):
            st.session_state.current_page = 'run_all'
            st.rerun()

    with col2:
        if st.button("📊 View Excel Files", key="dash_view", use_container_width=True):
            st.session_state.current_page = 'view_data'
            st.rerun()

    with col3:
        if st.button("🗑️ Reset System", key="dash_reset", use_container_width=True):
            st.session_state.current_page = 'reset'
            st.rerun()


def render_agent1_page():
    """Render Agent 1 execution page with form"""
    st.markdown("## 📧 Agent 1: Email Extractor")

    st.info("""
    **What this agent does:**
    - Connects to Gmail API
    - Searches for emails matching your criteria
    - Extracts GitHub URLs from email bodies
    - Creates Excel1.xlsx with extracted data
    """)

    # Render form
    params = render_agent1_form()

    if params:
        st.markdown("---")
        st.markdown("### 🔄 Execution")

        with st.spinner("Executing Agent 1..."):
            # Show progress
            progress_bar = st.progress(0)
            status_text = st.empty()

            status_text.text("Authenticating with Gmail...")
            progress_bar.progress(20)
            time.sleep(0.5)

            status_text.text("Searching emails...")
            progress_bar.progress(40)

            # Execute agent
            success = _execute_agent1(params)

            progress_bar.progress(100)

            if success:
                # Show results
                excel_path = get_excel_file('Excel1.xlsx')
                if excel_path.exists():
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    row_count = ws.max_row - 1
                    ready_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[6] == 'Ready')

                    st.success(f"✅ Agent 1 completed! Found **{row_count}** emails (**{ready_count}** with GitHub URLs)")

                    render_excel_viewer(excel_path, "Excel1.xlsx - Email Extraction Results")

                    # Next step suggestion
                    if ready_count > 0:
                        st.info("💡 **Next Step:** Run Agent 2 to analyze the repositories")
                        if st.button("▶️ Go to Agent 2", use_container_width=True):
                            st.session_state.current_page = 'agent2'
                            st.rerun()
                    else:
                        st.warning("⚠️ No emails with GitHub URLs found. Try adjusting your search parameters.")
            else:
                st.error("❌ Agent 1 failed. Check the logs for details.")


def render_agent2_page():
    """Render Agent 2 execution page"""
    st.markdown("## 🔬 Agent 2: Repository Analyzer")

    st.info("""
    **What this agent does:**
    - Reads Excel1.xlsx for GitHub URLs
    - Clones repositories (5 concurrent workers)
    - Analyzes Python code quality
    - Calculates grades based on line limits
    - Creates Excel2.xlsx with analysis results
    """)

    # Check prerequisite
    excel1 = get_excel_file('Excel1.xlsx')
    if not excel1.exists():
        st.warning("⚠️ Excel1.xlsx not found. Please run Agent 1 first.")
        if st.button("▶️ Go to Agent 1", use_container_width=True):
            st.session_state.current_page = 'agent1'
            st.rerun()
        return

    if st.button("🚀 Run Agent 2", use_container_width=True, type="primary"):
        st.markdown("---")
        st.markdown("### 🔄 Execution")

        with st.spinner("Executing Agent 2..."):
            progress_bar = st.progress(0)
            status_text = st.empty()

            status_text.text("Reading Excel1.xlsx...")
            progress_bar.progress(10)
            time.sleep(0.5)

            status_text.text("Cloning repositories (5 workers)...")
            progress_bar.progress(30)

            success = _execute_agent2()

            progress_bar.progress(100)

            if success:
                excel_path = get_excel_file('Excel2.xlsx')
                if excel_path.exists():
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    row_count = ws.max_row - 1

                    st.success(f"✅ Agent 2 completed! Analyzed **{row_count}** repositories")

                    render_excel_viewer(excel_path, "Excel2.xlsx - Repository Analysis")

                    if row_count > 0:
                        st.info("💡 **Next Step:** Run Agent 3 to generate AI feedback")
                        if st.button("▶️ Go to Agent 3", use_container_width=True):
                            st.session_state.current_page = 'agent3'
                            st.rerun()
            else:
                st.error("❌ Agent 2 failed. Check the logs for details.")


def render_agent3_page():
    """Render Agent 3 execution page"""
    st.markdown("## 🤖 Agent 3: AI Feedback Generator")

    st.info("""
    **What this agent does:**
    - Reads Excel2.xlsx for grades
    - Selects persona based on grade range:
      - 90-100: Trump (Enthusiastic)
      - 70-90: Shahar Hason (Comedian)
      - 55-70: Bruce Lee (Philosopher)
      - 0-55: Dudi Amsalem (Direct)
    - Generates personalized AI feedback
    - Creates Excel3.xlsx with feedback
    """)

    excel2 = get_excel_file('Excel2.xlsx')
    if not excel2.exists():
        st.warning("⚠️ Excel2.xlsx not found. Please run Agent 2 first.")
        if st.button("▶️ Go to Agent 2", use_container_width=True):
            st.session_state.current_page = 'agent2'
            st.rerun()
        return

    if st.button("🚀 Run Agent 3", use_container_width=True, type="primary"):
        st.markdown("---")
        st.markdown("### 🔄 Execution")

        with st.spinner("Executing Agent 3..."):
            progress_bar = st.progress(0)
            status_text = st.empty()

            status_text.text("Reading Excel2.xlsx...")
            progress_bar.progress(10)
            time.sleep(0.5)

            status_text.text("Generating AI feedback...")
            progress_bar.progress(40)

            success = _execute_agent3()

            progress_bar.progress(100)

            if success:
                excel_path = get_excel_file('Excel3.xlsx')
                if excel_path.exists():
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    row_count = ws.max_row - 1

                    st.success(f"✅ Agent 3 completed! Generated **{row_count}** feedback messages")

                    render_excel_viewer(excel_path, "Excel3.xlsx - AI Feedback")

                    if row_count > 0:
                        st.info("💡 **Next Step:** Run Agent 4 to create Gmail drafts")
                        if st.button("▶️ Go to Agent 4", use_container_width=True):
                            st.session_state.current_page = 'agent4'
                            st.rerun()
            else:
                st.error("❌ Agent 3 failed. Check the logs for details.")


def render_agent4_page():
    """Render Agent 4 execution page"""
    st.markdown("## ✉️ Agent 4: Draft Creator")

    st.info("""
    **What this agent does:**
    - Reads Excel3.xlsx for feedback
    - Joins with Excel1.xlsx for email data
    - Creates Gmail draft replies
    - Uses student mapping for personalization
    - Drafts are NOT sent automatically
    """)

    excel3 = get_excel_file('Excel3.xlsx')
    if not excel3.exists():
        st.warning("⚠️ Excel3.xlsx not found. Please run Agent 3 first.")
        if st.button("▶️ Go to Agent 3", use_container_width=True):
            st.session_state.current_page = 'agent3'
            st.rerun()
        return

    if st.button("🚀 Run Agent 4", use_container_width=True, type="primary"):
        st.markdown("---")
        st.markdown("### 🔄 Execution")

        with st.spinner("Executing Agent 4..."):
            progress_bar = st.progress(0)
            status_text = st.empty()

            status_text.text("Reading Excel3.xlsx...")
            progress_bar.progress(20)
            time.sleep(0.5)

            status_text.text("Creating Gmail drafts...")
            progress_bar.progress(60)

            success = _execute_agent4({})

            progress_bar.progress(100)

            if success:
                st.success("✅ Agent 4 completed successfully! Gmail drafts created.")
                st.balloons()

                st.info("💡 **Done!** Check your Gmail Drafts folder to review and send.")
            else:
                st.error("❌ Agent 4 failed. Check the logs for details.")


def render_run_all_page():
    """Render Run All Agents page"""
    st.markdown("## 🚀 Run All Agents")

    st.warning("""
    **This will execute all 4 agents sequentially:**
    1. Agent 1: Email Extractor
    2. Agent 2: Repository Analyzer
    3. Agent 3: AI Feedback Generator
    4. Agent 4: Draft Creator
    """)

    # Get Agent 1 parameters
    params = render_agent1_form()

    if params:
        st.markdown("---")
        st.markdown("### 🔄 Pipeline Execution")

        # Show search parameters
        st.info(f"""
        **Agent 1 Search Parameters:**
        - Subject: `{params.get('email_subject') or '(all subjects)'}`
        - Sender: `{params.get('sender_email') or '(all senders)'}`
        - Max Emails: `{params.get('max_emails', 20)}`
        """)

        # Execute all agents
        agents = [
            ("Agent 1: Email Extractor", _execute_agent1, params),
            ("Agent 2: Repository Analyzer", _execute_agent2, {}),
            ("Agent 3: AI Feedback Generator", _execute_agent3, {}),
            ("Agent 4: Draft Creator", _execute_agent4, {})
        ]

        overall_progress = st.progress(0)
        results_summary = []

        for i, (name, func, agent_params) in enumerate(agents):
            st.markdown(f"#### {name}")

            result_placeholder = st.empty()

            with st.spinner(f"Running {name}..."):
                # Only Agent 1 takes parameters
                if i == 0:
                    success = func(agent_params)
                else:
                    success = func()

                # Show results based on agent
                if success:
                    if i == 0:  # Agent 1
                        excel1 = get_excel_file('Excel1.xlsx')
                        if excel1.exists():
                            import openpyxl
                            wb = openpyxl.load_workbook(excel1)
                            ws = wb.active
                            row_count = ws.max_row - 1
                            ready_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[6] == 'Ready')
                            result_placeholder.success(f"✅ {name} completed! Found **{row_count}** emails (**{ready_count}** with GitHub URLs)")
                            results_summary.append(f"**Agent 1:** {row_count} emails ({ready_count} ready)")
                    elif i == 1:  # Agent 2
                        excel2 = get_excel_file('Excel2.xlsx')
                        if excel2.exists():
                            import openpyxl
                            wb = openpyxl.load_workbook(excel2)
                            ws = wb.active
                            row_count = ws.max_row - 1
                            result_placeholder.success(f"✅ {name} completed! Analyzed **{row_count}** repositories")
                            results_summary.append(f"**Agent 2:** {row_count} repos analyzed")
                    elif i == 2:  # Agent 3
                        excel3 = get_excel_file('Excel3.xlsx')
                        if excel3.exists():
                            import openpyxl
                            wb = openpyxl.load_workbook(excel3)
                            ws = wb.active
                            row_count = ws.max_row - 1
                            result_placeholder.success(f"✅ {name} completed! Generated **{row_count}** feedback messages")
                            results_summary.append(f"**Agent 3:** {row_count} feedback generated")
                    elif i == 3:  # Agent 4
                        result_placeholder.success(f"✅ {name} completed! Gmail drafts created")
                        results_summary.append("**Agent 4:** Drafts created")
                else:
                    result_placeholder.error(f"❌ {name} failed!")
                    st.stop()

            overall_progress.progress((i + 1) / len(agents))
            time.sleep(0.5)

        st.success("🎉 All agents completed successfully!")
        st.balloons()

        # Show summary
        st.markdown("### 📊 Pipeline Summary")
        for summary in results_summary:
            st.markdown(f"- {summary}")

        st.info("💡 Check the 'View Excel Files' page to see detailed results.")


def render_view_data_page():
    """Render Excel files viewer page"""
    st.markdown("## 📊 View Excel Files")

    tab1, tab2, tab3 = st.tabs(["📧 Excel1: Emails", "🔬 Excel2: Analysis", "🤖 Excel3: Feedback"])

    with tab1:
        excel1 = get_excel_file('Excel1.xlsx')
        render_excel_viewer(excel1, "Excel1.xlsx - Email Extraction Data")

    with tab2:
        excel2 = get_excel_file('Excel2.xlsx')
        render_excel_viewer(excel2, "Excel2.xlsx - Repository Analysis")

    with tab3:
        excel3 = get_excel_file('Excel3.xlsx')
        render_excel_viewer(excel3, "Excel3.xlsx - AI Feedback")


def render_status_page():
    """Render system status page"""
    st.markdown("## ℹ️ System Status")

    status_data = render_system_status()
    render_status_cards(status_data)

    st.markdown("---")

    # Detailed status
    st.markdown("### 📁 File Details")

    for name, data in status_data['excel_files'].items():
        with st.expander(f"📄 {name}"):
            if data['exists']:
                st.write(f"✅ **Status:** Exists")
                st.write(f"📊 **Rows:** {data['rows']}")
                st.write(f"🕒 **Last Modified:** {data['modified'].strftime('%Y-%m-%d %H:%M:%S')}")
            else:
                st.write(f"❌ **Status:** Not found")


def reset_system():
    """Reset the system by deleting Excel files and temp repos"""
    import shutil

    # Delete Excel files
    excel_files = ['Excel1.xlsx', 'Excel2.xlsx', 'Excel3.xlsx']
    excel_dir = get_excel_dir()

    for excel_file in excel_files:
        excel_path = excel_dir / excel_file
        if excel_path.exists():
            excel_path.unlink()

    # Clean temp repos
    clean_temp_repos()


def render_reset_page():
    """Render system reset page"""
    st.markdown("## 🗑️ Reset System")

    st.warning("""
    **⚠️ WARNING: This will delete:**
    - All Excel files (Excel1.xlsx, Excel2.xlsx, Excel3.xlsx)
    - All cloned repositories in temp/repos/
    - Logs remain intact

    This action cannot be undone!
    """)

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        confirm = st.text_input("Type 'RESET' to confirm:", key="reset_confirm")

        if st.button("🗑️ Reset System", use_container_width=True, type="primary"):
            if confirm == "RESET":
                with st.spinner("Resetting system..."):
                    reset_system()
                    st.success("✅ System reset successfully!")
                    time.sleep(1)
                    st.session_state.current_page = 'dashboard'
                    st.rerun()
            else:
                st.error("❌ Please type 'RESET' to confirm")


if __name__ == "__main__":
    main()
