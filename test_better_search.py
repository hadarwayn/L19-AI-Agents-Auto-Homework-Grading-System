"""
Test with a more specific search term
"""
from src.ui.agent_runner import _execute_agent1

# Try with "Homework - L" which is more specific
email_subject = "Homework - L"
sender_email = None
max_emails = 20

print("="*70)
print("Testing with more specific search term")
print("="*70)
print(f"\nEmail Subject: {email_subject}")
print(f"Max Emails: {max_emails}\n")

# Build query using current logic
words = email_subject.split()
if len(words) <= 3:
    query = f'subject:{email_subject}'
else:
    key_words = ' '.join(words[:3])
    query = f'subject:{key_words}'

print(f"Gmail Query: {query}")
print(f"Post-Filter: '{email_subject}' as substring\n")

params = {
    'email_subject': email_subject,
    'sender_email': sender_email,
    'max_emails': max_emails,
    'query': query
}

print("Executing Agent 1...")
print("="*70 + "\n")

success = _execute_agent1(params)

if success:
    import openpyxl
    from src.utils.paths import get_excel_file

    excel_path = get_excel_file('Excel1.xlsx')
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        row_count = ws.max_row - 1
        print(f"\n{'='*70}")
        print(f"RESULT: {row_count} emails found")
        print(f"{'='*70}\n")

        if row_count > 0:
            print("Email subjects found:")
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                email_id, received_time, subject, sender, github_url, thread_id, status = row
                print(f"  {i}. {subject}")

            if row_count == 8:
                print(f"\nSUCCESS: Found all 8 homework emails (L11-L18)")
            else:
                print(f"\nFound {row_count} emails")
