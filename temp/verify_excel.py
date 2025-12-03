import openpyxl

wb = openpyxl.load_workbook('results/excel/Excel1.xlsx')
ws = wb.active

print(f"Total rows: {ws.max_row}")
print(f"\nHeaders: {[cell.value for cell in ws[1]]}")
print(f"\nExtracted Emails:")
print("="*100)

ready_count = 0
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
    email_id, received_time, subject, sender, hashed, github_url, thread_id, status = row
    print(f"{idx}. Subject: {subject[:70]}...")
    print(f"   GitHub: {github_url if github_url else '(none)'}")
    print(f"   Status: {status}")
    print()
    if status == "Ready":
        ready_count += 1

print("="*100)
print(f"\n✅ Total emails extracted: {ws.max_row - 1}")
print(f"✅ Emails with GitHub URLs (Ready): {ready_count}")
print(f"⚠️  Emails without GitHub URLs: {ws.max_row - 1 - ready_count}")
