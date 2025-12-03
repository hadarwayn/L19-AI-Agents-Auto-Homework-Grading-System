import openpyxl

wb = openpyxl.load_workbook('results/excel/Excel2.xlsx')
ws = wb.active

print(f"Total rows: {ws.max_row}")
print(f"\nHeaders: {[cell.value for cell in ws[1]]}")
print(f"\nRepository Analysis Results:")
print("="*100)

for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
    email_id, github_url, total_files, total_lines, compliant_lines, grade, status = row
    print(f"{idx}. Email ID: {email_id[:16]}...")
    print(f"   GitHub: {github_url[:70]}...")
    print(f"   Files: {total_files}, Lines: {total_lines}, Compliant: {compliant_lines}")
    print(f"   Grade: {grade}%, Status: {status}")
    print()

print("="*100)
print(f"\nTotal repositories analyzed: {ws.max_row - 1}")
successful = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[6] == "Ready")
print(f"Successful: {successful}")
print(f"Failed: {ws.max_row - 1 - successful}")
