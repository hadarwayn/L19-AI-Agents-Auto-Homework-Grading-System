"""
Test Agent 1 with the user's search term
"""
from src.ui.agent_runner import _prompt_agent1_parameters, _execute_agent1

print("Testing with: 'AI Development Expert course'")
print("="*70)

# Simulate user input
email_subject = "AI Development Expert course"
sender_email = None
max_emails = 10

# Build query using new logic
words = email_subject.split()
if len(words) == 1:
    query = f'subject:{email_subject}'
elif len(words) > 6:
    key_words = ' '.join(words[:5])
    query = f'subject:"{key_words}"'
else:
    query = f'subject:"{email_subject}"'

print(f"\nEmail Subject: {email_subject}")
print(f"Gmail Query: {query}")
print(f"Max Emails: {max_emails}\n")

params = {
    'email_subject': email_subject,
    'sender_email': sender_email,
    'max_emails': max_emails,
    'query': query
}

print("Executing Agent 1...")
print("="*70 + "\n")

success = _execute_agent1(params)

if success:
    import openpyxl
    from src.utils.paths import get_excel_file

    excel_path = get_excel_file('Excel1.xlsx')
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        row_count = ws.max_row - 1
        print(f"\n{'='*70}")
        print(f"RESULT: {row_count} emails found")
        print(f"{'='*70}\n")

        if row_count > 0:
            print("Email subjects:")
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                email_id, received_time, subject, sender, github_url, thread_id, status = row
                print(f"{i}. {subject}")
