"""
Debug Agent 1 with specific parameters
"""
import sys
from pathlib import Path

# Test parameters
params = {
    'email_subject': 'AI Development Expert course - Homework - L',
    'sender_email': None,
    'max_emails': 20,
    'query': 'subject:AI Development Expert course - Homework - L'
}

print("="*70)
print("TESTING AGENT 1 WITH DEBUG")
print("="*70)
print(f"Email Subject: {params['email_subject']}")
print(f"Sender Email: {params['sender_email']}")
print(f"Max Emails: {params['max_emails']}")
print(f"Gmail Query: {params['query']}")
print("="*70 + "\n")

# Import and run Agent 1
from src.ui.agent_runner import _execute_agent1

print("\nExecuting Agent 1...\n")
success = _execute_agent1(params)

if success:
    print("\n[+] Agent 1 completed successfully!")

    # Check the Excel file
    import openpyxl
    from src.utils.paths import get_excel_file

    excel_path = get_excel_file('Excel1.xlsx')
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        row_count = ws.max_row - 1  # Subtract header
        print(f"\n[+] Excel1.xlsx created with {row_count} emails")

        print("\nFirst 10 email subjects:")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 1):
            email_id, received_time, subject, sender, github_url, thread_id, status = row
            print(f"{i}. Subject: {subject}")
    else:
        print("\n[!] Excel1.xlsx not found!")
else:
    print("\n[!] Agent 1 failed!")
